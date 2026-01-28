# E:\work\backend\orders\views.py
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
import json

# 订单Excel导出接口（必须确保函数名是 export_order_excel）
@csrf_exempt
@require_POST
def export_order_excel(request):
    try:
        # 1. 接收前端传递的参数
        params = json.loads(request.body)
        selected_date = params.get('selectedDate')
        selected_canteen = params.get('selectedCanteen')
        selected_category = params.get('selectedCategory')
        export_type = params.get('exportType')
        excel_data = params.get('excelData')

        if not excel_data or not export_type or not selected_date:
            return HttpResponse('缺少必要参数', status=400)

        # 2. 创建Excel工作簿
        wb = Workbook()
        ws = wb.active

        # 3. 按导出类型生成Excel
        if export_type == 'supplier':
            # 供货商导出逻辑
            ws.title = '供货商数据'
            headers = ['学校名称', '食堂名称', '配送日期', '食材类别', '食材名称', '规格', '单位', '数量', '订单备注']
            ws.append(headers)
            for row in excel_data:
                final_qty = row.get('实际增减补') or row.get('原始数量')
                ws.append([
                    row.get('学校名称', ''),
                    row.get('食堂名称', ''),
                    row.get('配送日期', ''),
                    row.get('食材类别', ''),
                    row.get('食材名称', ''),
                    row.get('规格', ''),
                    row.get('单位', ''),
                    final_qty,
                    row.get('订单备注', '')
                ])
            # 设置列宽
            col_widths = [15, 15, 15, 15, 18, 12, 8, 10, 25]
            for i, width in enumerate(col_widths):
                col_letter = get_column_letter(i+1)
                ws.column_dimensions[col_letter].width = width

        elif export_type == 'follower':
            # 跟车单导出逻辑
            ws.title = '跟车单数据'
            # 标题行（合并7列）
            title = f'【{selected_canteen}】跟车单（配送日期：{selected_date}）'
            ws.merge_cells('A1:G1')
            ws['A1'] = title
            ws['A1'].font = Font(bold=True, size=12)
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            # 表头行
            headers = ['食材名称', '标记', '备注', '规格', '单位', '下单数量', '实际数量']
            ws.append(headers)
            # 填充数据
            for row in excel_data:
                actual_qty = row.get('实际增减补') or row.get('原始数量')
                ws.append([
                    row.get('食材名称', ''),
                    '',  # 标记列
                    row.get('订单备注', ''),
                    row.get('规格', ''),
                    row.get('单位', ''),
                    row.get('原始数量', ''),
                    actual_qty
                ])
            # 设置列宽
            col_widths = [18, 8, 25, 12, 8, 12, 12]
            for i, width in enumerate(col_widths):
                col_letter = get_column_letter(i+1)
                ws.column_dimensions[col_letter].width = width
            # 下单数量/实际数量列居中
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=7):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')

        # 4. 配置响应头，返回文件流
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # 生成文件名
        if export_type == 'supplier':
            filename = f'供货商单_配送日期_{selected_date}_食材类别_{selected_category}.xlsx'
        else:
            filename = f'跟车单_配送日期_{selected_date}_食堂名称_{selected_canteen}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)

        return response

    except Exception as e:
        print(f'导出Excel失败：{str(e)}')
        return HttpResponse(f'导出失败：{str(e)}', status=500)